import openpyxl
import re
import json
import os

def clean_addr(addr):
    if not addr:
        return ""
    # Extract the address portion like 12.00.03 from "12.00.03 (06)" or "12.00.03"
    match = re.match(r'^(\d+\.\d+\.\d+)', str(addr).strip())
    if match:
        return match.group(1)
    return str(addr).strip()

def main():
    excel_file = "bacnet configuration.xlsx"
    if not os.path.exists(excel_file):
        print(f"Erro: Arquivo '{excel_file}' não encontrado no diretório atual.")
        return
        
    print(f"Lendo '{excel_file}'...")
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    # Ler Planilha1 para mapear Endereço -> Object ID (Device ID)
    sheet1 = wb["Planilha1"]
    device_objids = {} # addr -> object_id
    for r in range(13, sheet1.max_row + 1):
        cell_addr = sheet1.cell(row=r, column=3).value
        cell_objid = sheet1.cell(row=r, column=5).value
        cell_disp = sheet1.cell(row=r, column=2).value
        
        disp_str = str(cell_disp) if cell_disp else ""
        addr_str = str(cell_addr) if cell_addr else ""
        
        if "Unidade interior" in disp_str:
            addr = clean_addr(addr_str)
            if addr and cell_objid:
                device_objids[addr] = int(cell_objid)
                
    # Ler Planilha2 para mapear Endereço -> Nome Amigável
    sheet2 = wb["Planilha2"]
    devices = {}
    for r in range(2, sheet2.max_row + 1):
        cell_disp = sheet2.cell(row=r, column=2).value
        cell_addr = sheet2.cell(row=r, column=3).value
        cell_name = sheet2.cell(row=r, column=4).value
        
        disp_str = str(cell_disp) if cell_disp else ""
        addr_str = str(cell_addr) if cell_addr else ""
        name_str = str(cell_name) if cell_name else ""
        
        if "Unidade interior" in disp_str:
            addr = clean_addr(addr_str)
            if addr and name_str:
                objid = device_objids.get(addr)
                # Chave única por nome para busca rápida
                devices[name_str.upper()] = {
                    "name": name_str,
                    "address": addr,
                    "device_id": objid,
                    "long_name": disp_str,
                    "alias": name_str.lower()
                }
                
    print(f"Total de evaporadoras mapeadas: {len(devices)}")
    
    # Salvar em devices.json
    output_file = "devices.json"
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(devices, f, indent=4, ensure_ascii=False)
        
    print(f"Sucesso! Arquivo '{output_file}' gerado com sucesso.")

if __name__ == "__main__":
    main()
